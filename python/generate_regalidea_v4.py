import os
from openpyxl.styles import Protection
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

DARK_BLUE  = "203764"
MID_BLUE   = "305496"
YELLOW     = "FFF2CC"
LIGHT_BLUE = "D9E1F2"
AUTO_FILL  = "DDEEFF"
WHITE      = "FFFFFF"
BLACK      = "000000"
GRAY_TEXT  = "595959"

def fnt(bold=False, size=11, color=BLACK, name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color)
def fill(h):
    return PatternFill("solid", fgColor=h)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def box(color="BFBFBF", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def hdr_border():
    return Border(
        left=Side(style="thin", color=WHITE),
        right=Side(style="thin", color=WHITE),
        bottom=Side(style="medium", color="1F4E79"),
    )

wb = Workbook()

# ══════════════════════════════════════════════════════════════
# SHEET 1 — ANAGRAFICA_CLIENTI (hidden)
# ══════════════════════════════════════════════════════════════
ana = wb.active
ana.title = "ANAGRAFICA_CLIENTI"
ana.sheet_state = "hidden"
ana.sheet_properties.tabColor = "7F7F7F"

ana_cols = [
    ("NOME CLIENTE",         35),   # A
    ("P.IVA",                18),   # B
    ("CODE",                 35),   # C
    ("INDIRIZZO",            40),   # D
    ("TELEFONO",             16),   # E
    ("EMAIL",                30),   # F
    ("FAX",                  16),   # G
    ("CONDIZIONI PAGAMENTO", 28),   # H
    ("LISTINO / SCONTO STD", 22),   # I
]
for ci, (title, w) in enumerate(ana_cols, start=1):
    c = ana.cell(row=1, column=ci, value=title)
    c.font, c.fill, c.alignment = fnt(bold=True, size=10, color=WHITE), fill(DARK_BLUE), aln("center")
    ana.column_dimensions[chr(64+ci)].width = w

dummy_clients = [
    ("Rossi & Figli S.r.l.",         "IT01234567890", "C243545",
     "Via Roma 1, 10100 Torino",      "+39 011 1234567", "ordini@rossi.it",
     "+39 011 7654321",               "30 gg d.f.f.m.",   "Listino A -10%"),
    ("Bianchi Distribuzione S.p.A.",  "IT09876543210", "C948964959",
     "Corso Italia 55, 20100 Milano", "+39 02 9876543",  "acquisti@bianchi.it",
     "+39 02 3456789",                "60 gg d.f.",       "Listino B -15%"),
    ("Verdi Commercio S.n.c.",        "IT05555555500", "C385282",
     "Piazza Garibaldi 3, 40100 Bologna", "+39 051 555666", "info@verdi.it",
     "",                              "Rimessa diretta",  "Listino C -5%"),
]
for ri, row_data in enumerate(dummy_clients, start=2):
    for ci, val in enumerate(row_data, start=1):
        c = ana.cell(row=ri, column=ci, value=val)
        c.font, c.alignment = fnt(size=10), aln("left")

# ══════════════════════════════════════════════════════════════
# SHEET 1.5 — DESTINAZIONI (hidden)
# ══════════════════════════════════════════════════════════════
dest = wb.create_sheet("DESTINAZIONI")
dest.sheet_state = "hidden"
dest.sheet_properties.tabColor = "7F7F7F"

dest.column_dimensions["A"].width = 35
dest.column_dimensions["B"].width = 60

for ci, title in enumerate(["NOME CLIENTE", "INDIRIZZO CONSEGNA"], start=1):
    c = dest.cell(row=1, column=ci, value=title)
    c.font, c.fill, c.alignment = fnt(bold=True, size=10, color=WHITE), fill(DARK_BLUE), aln("center")

dummy_dest = [
    ("Rossi & Figli S.r.l.", "Via Roma 1, Torino (Sede Principale)"),
    ("Rossi & Figli S.r.l.", "Magazzino Esterno - Via Milano 45, Nichelino (TO)"),
    ("Bianchi Distribuzione S.p.A.", "Corso Italia 55, Milano"),
    ("Verdi Commercio S.n.c.", "Piazza Garibaldi 3, Bologna (Negozio Centro)"),
    ("Verdi Commercio S.n.c.", "Via dell'Industria 12, San Lazzaro (Magazzino)"),
]

for ri, (cliente, indirizzo) in enumerate(dummy_dest, start=2):
    dest.cell(row=ri, column=1, value=cliente).font = fnt(size=10)
    dest.cell(row=ri, column=2, value=indirizzo).font = fnt(size=10)

# ══════════════════════════════════════════════════════════════
# SHEET 2 — CATALOGO (hidden)
# ══════════════════════════════════════════════════════════════
cat = wb.create_sheet("CATALOGO")
cat.sheet_state = "hidden"
cat.sheet_properties.tabColor = "7F7F7F"

cat_cols = [("NOME CLIENTE", 35), ("DESCRIZIONE ARTICOLO", 52), ("CODICE FORNITORE", 20), ("CODICE ARTICOLO", 20)]
for ci, (title, w) in enumerate(cat_cols, start=1):
    c = cat.cell(row=1, column=ci, value=title)
    c.font, c.fill, c.alignment = fnt(bold=True, size=10, color=WHITE), fill(MID_BLUE), aln("center")
    cat.column_dimensions[chr(64+ci)].width = w

dummy_catalogue = [
    ("Rossi & Figli S.r.l.",         "Candela Profumata Vaniglia 200g",      "60001", "INT-A01"),
    ("Rossi & Figli S.r.l.",         "Set Regalo Lusso Natale 2025",         "60002", "INT-A02"),
    ("Rossi & Figli S.r.l.",         "Diffusore Ambiente Bergamotto 100ml",  "60003", "INT-A03"),
    ("Bianchi Distribuzione S.p.A.", "Portafotografie Legno Naturale 15x20", "60101", "INT-B01"),
    ("Bianchi Distribuzione S.p.A.", "Cornice Argento Premium 20x30",        "60102", "INT-B02"),
    ("Bianchi Distribuzione S.p.A.", "Album Foto Pelle 100 Foto",            "60103", "INT-B03"),
    ("Verdi Commercio S.n.c.",       "Vassoio Decorativo Metallo Oro 40cm",  "60201", "INT-C01"),
    ("Verdi Commercio S.n.c.",       "Cesto Regalo Vimini Grande",           "60202", "INT-C02"),
    ("Verdi Commercio S.n.c.",       "Scatola Regalo Rigida Rossa 30x20x10", "60203", "INT-C03"),
]

for ri, (cliente, desc, cod_forn, cod_art) in enumerate(dummy_catalogue, start=2):
    cat.cell(row=ri, column=1, value=cliente).font = fnt(size=10)
    cat.cell(row=ri, column=2, value=desc).font    = fnt(size=10)
    cat.cell(row=ri, column=3, value=cod_forn).font = fnt(size=10, bold=True)
    cat.cell(row=ri, column=4, value=cod_art).font  = fnt(size=10, bold=True)

# ══════════════════════════════════════════════════════════════
# SHEET 3 — HELPER (hidden)
# ══════════════════════════════════════════════════════════════
hlp = wb.create_sheet("HELPER")
hlp.sheet_state = "hidden"
hlp.sheet_properties.tabColor = "7F7F7F"
hlp.column_dimensions["A"].width = 52
hlp.column_dimensions["B"].width = 22

for ci, title in enumerate(["DESCRIZIONE (filtrata)", "CODICE FORNITORE (filtrato)"], start=1):
    c = hlp.cell(row=1, column=ci, value=title)
    c.font, c.fill, c.alignment = fnt(bold=True, size=9, color=WHITE), fill(DARK_BLUE), aln("center")

hlp.cell(row=2, column=1, value='=_xlfn.FILTER(CATALOGO!B2:B10000, CATALOGO!A2:A10000=ORDINE!$H$4, "NESSUN ARTICOLO TROVATO")')
hlp.cell(row=2, column=2, value='=_xlfn.FILTER(CATALOGO!C2:C10000, CATALOGO!A2:A10000=ORDINE!$H$4, "")')
hlp.cell(row=2, column=3, value='=_xlfn.FILTER(CATALOGO!D2:D10000, CATALOGO!A2:A10000=ORDINE!$H$4, "")')

hlp.column_dimensions["E"].width = 60
c = hlp.cell(row=1, column=5, value="DESTINAZIONI (filtrate)")
c.font, c.fill, c.alignment = fnt(bold=True, size=9, color=WHITE), fill(DARK_BLUE), aln("center")
hlp.cell(row=2, column=5, value='=_xlfn.FILTER(DESTINAZIONI!B2:B10000, DESTINAZIONI!A2:A10000=ORDINE!$H$4, "")')

# ── NEW: GENERATE NUMBERS 1 TO 200 FOR QUANTITY DROPDOWN ──
hlp.column_dimensions["G"].width = 20
c = hlp.cell(row=1, column=7, value="QUANTITÀ AMMESSE")
c.font, c.fill, c.alignment = fnt(bold=True, size=9, color=WHITE), fill(DARK_BLUE), aln("center")

# Write numbers 1 to 200 into column G
for i in range(1, 201):
    hlp.cell(row=i+1, column=7, value=i)

# ══════════════════════════════════════════════════════════════
# SHEET 4 — ORDINE (visible)
# ══════════════════════════════════════════════════════════════
ws = wb.create_sheet("ORDINE")
ws.sheet_properties.tabColor = DARK_BLUE

col_widths = {
    "A": 20, "B": 20, "C": 20, "D": 40, "E": 8, "F": 22, "G": 22, "H": 22, "I": 22
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w
for r in range(1, 170):
    ws.row_dimensions[r].height = 18

ws.merge_cells("A1:I1")
c = ws["A1"]
c.value = "DRINK INTERNATIONAL SPA — ORDINE FORNITORE"
c.font, c.fill, c.alignment = fnt(bold=True, size=12, color=WHITE), fill(DARK_BLUE), aln("center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:I2")
c = ws["A2"]
c.value = "Seleziona il CLIENTE (H4) → dati anagrafici e catalogo si aggiornano. Poi scegli DESCRIZIONE → Codice Fornitore si compila da solo."
c.font, c.fill, c.alignment = fnt(size=10, color=WHITE), fill(MID_BLUE), aln("center")
ws.row_dimensions[3].height = 6

def label(row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font, c.alignment = fnt(bold=True, color=DARK_BLUE), aln("right")

def yellow_merged(addr_from, addr_to, row):
    for col_letter in range(ord(addr_from), ord(addr_to) + 1):
        cell = ws[f"{chr(col_letter)}{row}"]
        cell.protection = Protection(locked=False)
    ws.merge_cells(f"{addr_from}{row}:{addr_to}{row}")
    c = ws[f"{addr_from}{row}"]
    c.fill, c.font, c.alignment, c.border = fill(YELLOW), fnt(size=11), aln("left"), box("E6C800")
    return c

def auto_merged(addr_from, addr_to, row, formula):
    ws.merge_cells(f"{addr_from}{row}:{addr_to}{row}")
    c = ws[f"{addr_from}{row}"]
    c.value  = formula
    c.fill   = fill(AUTO_FILL)
    c.font   = fnt(size=11, color=GRAY_TEXT)
    c.border = box("9DC3E6")
    return c

def vlook(col_n):
    return f'=IFERROR(VLOOKUP($H$4,ANAGRAFICA_CLIENTI!$A:$I,{col_n},0),"")'

label(4,  1, "N. DOCUMENTO:");    yellow_merged("B","E", 4)
label(5,  1, "DATA EMISSIONE:");  c5=yellow_merged("B","E",5); c5.number_format="DD/MM/YYYY"
label(6,  1, "DATA CONSEGNA:");   c6=yellow_merged("B","E",6); c6.number_format="DD/MM/YYYY"
label(7,  1, "DESTINAZIONE:");    yellow_merged("B","E", 7)
label(8,  1, "NOTE ORDINE:");     yellow_merged("B","E", 8)

label(4,  6, "NOME CLIENTE:");    yellow_merged("H","I", 4) 
label(5,  6, "P.IVA:");           auto_merged("H","I", 5,  vlook(2))
label(6,  6, "CODE:");            auto_merged("H","I", 6,  vlook(3))
label(7,  6, "INDIRIZZO:");       auto_merged("H","I", 7,  vlook(4))
label(8,  6, "TELEFONO:");        auto_merged("H","I", 8,  vlook(5))
label(9,  6, "EMAIL:");           auto_merged("H","I", 9,  vlook(6))
label(10, 6, "FAX:");             auto_merged("H","I", 10, vlook(7))

ws.row_dimensions[11].height = 6

label(12, 1, "COND. PAGAMENTO:"); auto_merged("B","E", 12, vlook(8))
label(12, 6, "LISTINO / SCONTO:"); auto_merged("H","I", 12, vlook(9))

for row, lbl, val in [(13, "FORNITORE:", "DRINK INTERNATIONAL SPA"), (14, "P.IVA FORN.:", "IT00926410010")]:
    label(row, 6, lbl)
    ws.merge_cells(f"H{row}:I{row}")
    c = ws[f"H{row}"]
    c.value, c.font, c.alignment = val, fnt(size=11), aln("left")

ws.row_dimensions[15].height = 6

label(16, 1, "TOTALE QUANTITÀ:")
ws.merge_cells("B16:E16")
tc = ws["B16"]
tc.value, tc.fill, tc.font, tc.alignment = "=SUM(F19:F500)", fill(LIGHT_BLUE), fnt(size=11), aln("center")
tc.border, tc.number_format = box("9DC3E6"), "#,##0"

label(16, 6, "TOTALE IMPORTO:")
ws.merge_cells("H16:I16")
ti = ws["H16"]
ti.value, ti.fill = "=SUM(I19:I500)", fill(LIGHT_BLUE)
ti.font, ti.alignment, ti.border, ti.number_format = fnt(bold=True, size=11, color=DARK_BLUE), aln("center"), box("9DC3E6"), '€ #,##0.00'

ws.row_dimensions[17].height = 6

ws.row_dimensions[18].height = 26
col_headers = ["#", "CODICE FORNITORE", "CODICE ARTICOLO", "DESCRIZIONE", "U.M.", "QTÀ", "PREZZO LORDO €", "SCONTO %", "IMPORTO NETTO €"]
for ci, title in enumerate(col_headers, start=1):
    c = ws.cell(row=18, column=ci, value=title)
    c.font, c.fill, c.alignment, c.border = fnt(bold=True, size=11, color=WHITE), fill(DARK_BLUE), aln("center", wrap=True), hdr_border()

DATA_START, DATA_END = 19, 500
fw = fill(WHITE); fs = fill("EBF1F8"); fy = fill(YELLOW); fl = fill(LIGHT_BLUE); fa = fill(AUTO_FILL)
unlocked = Protection(locked=False)

for row in range(DATA_START, DATA_END + 1):
    n    = row - DATA_START + 1
    rf   = fs if n % 2 == 0 else fw
    thin = box("D0D8E8")
    ybox = box("E6C800")

    c = ws.cell(row=row, column=1, value=n)
    c.font, c.fill, c.alignment, c.border = fnt(bold=True, color=DARK_BLUE, size=10), rf, aln("center"), thin

    b = ws.cell(row=row, column=2)
    b.value = f'=IFERROR(IF(D{row}="","",VLOOKUP(D{row},HELPER!$A:$C,2,0)&""),"")'
    b.fill, b.font, b.alignment, b.border = fa, fnt(size=11, color=GRAY_TEXT), aln("center"), thin

    c2 = ws.cell(row=row, column=3)
    c2.value = f'=IFERROR(IF(D{row}="","",VLOOKUP(D{row},HELPER!$A:$C,3,0)&""),"")'
    c2.fill, c2.font, c2.alignment, c2.border = fa, fnt(size=11, color=GRAY_TEXT), aln("center"), thin

    d = ws.cell(row=row, column=4)
    d.fill, d.font, d.alignment, d.border = fy, fnt(size=11), aln("left", wrap=True), ybox
    d.protection = unlocked 

    e = ws.cell(row=row, column=5)
    e.fill, e.font, e.alignment, e.border = fy, fnt(size=11), aln("center"), ybox
    e.protection = unlocked 

    f_c = ws.cell(row=row, column=6)
    f_c.fill, f_c.font, f_c.alignment, f_c.border = fy, fnt(size=11), aln("center"), ybox
    f_c.number_format = "#,##0"
    f_c.protection = unlocked 

    g = ws.cell(row=row, column=7)
    g.fill, g.font, g.alignment, g.border = fy, fnt(size=11), aln("right"), ybox
    g.number_format = '€ #,##0.00'
    g.protection = unlocked 

    h = ws.cell(row=row, column=8)
    h.fill, h.font, h.alignment, h.border = fy, fnt(size=11), aln("center"), ybox
    h.number_format = '0.00%'
    h.protection = unlocked 

    i_c = ws.cell(row=row, column=9)
    i_c.value = f'=IF(F{row}="","",ROUND(F{row}*G{row}*(1-H{row}),2))'
    i_c.fill, i_c.font, i_c.alignment, i_c.border = fl, fnt(bold=True, size=11, color=DARK_BLUE), aln("right"), box("9DC3E6")
    i_c.number_format = '€ #,##0.00'

# ══════════════════════════════════════════════════════════════
# DATA VALIDATIONS
# ══════════════════════════════════════════════════════════════

# 1. H4 — NOME CLIENTE
dv_cliente = DataValidation(
    type="list", formula1="=ANAGRAFICA_CLIENTI!$A$2:$A$10000",
    allow_blank=True, showDropDown=False, showErrorMessage=True,
    errorTitle="Cliente non trovato", error="Seleziona il cliente dal menu a tendina.",
    showInputMessage=True, promptTitle="Seleziona Cliente", prompt="Scegli il nome cliente."
)
dv_cliente.sqref = "H4"
ws.add_data_validation(dv_cliente)

# 2. Col D — DESCRIZIONE
dv_desc = DataValidation(
    type="list", formula1='=OFFSET(HELPER!$A$2, 0, 0, MAX(1, COUNTIF(HELPER!$A$2:$A$10000, "?*")), 1)',
    allow_blank=True, showDropDown=False, showErrorMessage=True,
    errorTitle="Articolo non disponibile", error="Seleziona prima il CLIENTE in H4.",
    showInputMessage=True, promptTitle="Seleziona Descrizione", prompt="Scegli la descrizione articolo."
)
dv_desc.sqref = f"D{DATA_START}:D{DATA_END}"
ws.add_data_validation(dv_desc)

# 3. B7 — DESTINAZIONE 
dv_dest = DataValidation(
    type="list", formula1='=OFFSET(HELPER!$E$2, 0, 0, MAX(1, COUNTIF(HELPER!$E$2:$E$10000, "?*")), 1)',
    allow_blank=True, showDropDown=False, showErrorMessage=True, 
    errorTitle="Selezione non valida", error="Devi selezionare un indirizzo di destinazione dal menu a tendina.",
    showInputMessage=True, promptTitle="Seleziona Punto Vendita", prompt="Scegli l'indirizzo di destinazione dal menu."
)
dv_dest.sqref = "B7:E7"
ws.add_data_validation(dv_dest)

# ── NEW: 4. Col F — QTÀ (Strictly lock to list 1-200) ─────────
dv_qta = DataValidation(
    type="list",
    formula1="=HELPER!$G$2:$G$201",  # Points to the new 1-200 list generated above
    allow_blank=True, 
    showDropDown=False,
    showErrorMessage=True,
    errorStyle="stop",  # This physically prevents manual typing
    errorTitle="Quantità non consentita",
    error="Seleziona una quantità valida (da 1 a 200) dal menu a tendina. L'inserimento manuale è bloccato.",
    showInputMessage=True,
    promptTitle="Seleziona Quantità",
    prompt="Clicca sulla freccia per scegliere la quantità (max 200).",
)
dv_qta.sqref = f"F{DATA_START}:F{DATA_END}"
ws.add_data_validation(dv_qta)

# ── Page setup ────────────────────────────────────────────────
ws.freeze_panes = "A19"
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize   = ws.PAPERSIZE_A4
ws.page_setup.fitToPage   = True
ws.page_setup.fitToWidth  = 1
ws.page_setup.fitToHeight = 0
ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
ws.print_title_rows = "1:18"
ws.protection.sheet = True 

# ══════════════════════════════════════════════════════════════
# SHEET 5 — ISTRUZIONI
# ══════════════════════════════════════════════════════════════
inst = wb.create_sheet("ISTRUZIONI")
inst.sheet_properties.tabColor = "70AD47"
inst.column_dimensions["A"].width = 2
inst.column_dimensions["B"].width = 100

inst.merge_cells("A1:B1")
c = inst["A1"]
c.value = "ISTRUZIONI PER LA COMPILAZIONE"
c.font, c.fill, c.alignment = fnt(bold=True, size=13, color=WHITE), fill(DARK_BLUE), aln("center")
inst.row_dimensions[1].height = 30

sections = [
    ("COME COMPILARE UN ORDINE", True),
    ("1.  Aprire il foglio ORDINE.", False),
    ("2.  Cliccare su H4 (NOME CLIENTE) e scegliere il cliente dal menu a tendina.", False),
    ("3.  Per ogni riga articolo: cliccare sulla colonna D (DESCRIZIONE) e scegliere dal menu.", False),
    ("4.  Inserire U.M., QTÀ, PREZZO LORDO e SCONTO %. L'importo netto è calcolato automaticamente.", False),
    ("", False),
    ("COME CARICARE I DATI PANTHERA", True),
    ("ANAGRAFICA_CLIENTI — incollare da riga 2.", False),
    ("CATALOGO — incollare da riga 2.", False),
    ("HELPER — non toccare mai questo foglio.", False),
    ("", False),
    ("LEGENDA COLORI", True),
    ("  🟡 GIALLO  — cella da compilare", False),
    ("  🔵 AZZURRO — cella auto-compilata", False),
    ("  🟦 BLU     — valore calcolato", False),
]

ri = 3
for text, is_title in sections:
    inst.row_dimensions[ri].height = 22 if is_title else 18
    inst.merge_cells(f"A{ri}:B{ri}")
    c = inst[f"A{ri}"]
    c.value = text
    if is_title:
        c.font, c.fill, c.alignment = fnt(bold=True, size=10, color=WHITE), fill(MID_BLUE), aln("center")
    else:
        c.font = fnt(size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ri += 1

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "DRINK_Smart_Template_NEW.xlsx")
wb.save(out)
print(f"✅  Saved → {out}")